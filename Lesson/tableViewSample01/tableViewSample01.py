import os
import sys
import dataclasses
import openpyxl

from typing import Any, List
from PySide2.QtCore import (
    Qt,
    QModelIndex,
    QAbstractTableModel
)

# For Sample
from PySide2 import QtWidgets, QtCore, QtGui
from PySide2.QtCore import *
from PySide2.QtGui import *
from PySide2.QtWidgets import *

path, fname = os.path.split(__file__)
os.chdir(path)


@dataclasses.dataclass
class Register:
    address: int
    bit: int
    size: int    
    def __init__(self, address, bit, size):
        self.address = address
        self.bit = bit
        self.size = size
    

# custom Data class
@dataclasses.dataclass
class GpioPort:
    name: str
    port: int
    signal: int
    direction: int
    value: int
    enable: int
    drive: int
    boost: int
    signal_port: Register 
    direction_port: Register 
    value_port: Register 
    enable_port: Register 
    drive_port: Register 
    boost_port: Register 
    values_signal: list

    def toList(self) -> list:
       return [self.name, self.port, self.signal, self.direction, 
            self.value, self.enable, self.drive, self.boost]

    @classmethod
    def toHeaderList(cls) -> List[str]:
        return ["Name", "Port", "Signal", "Direction", "Value", "Enable", "Drive", "Boost"]

    def validateAndSetName(self, name) -> bool:
        if str(name) != "":  # Non-empty string is OK.
            self.name = str(name)
            return True
        return False

    def validateAndSetPort(self, port) -> bool:
        try:
            val_int = int(port)
            if val_int >= 0:  # int, and greater than 0 or equal 0 is OK.
                self.port = val_int
                return True
        except ValueError:
            pass  # Non-integer value is invalid.
        return False

    def validateAndSetSignal(self, signal) -> bool:
        if signal in range(8):  # Only 5 countries
            self.signal = signal
            return True
        return False

    def validateAndSetDirection(self, direction) -> bool:
        if direction in (0, 1):
            self.direction = direction
            return True
        return False

    def validateAndSetValue(self, value) -> bool:
        if value in (0, 1):
            self.value = value
            return True
        return False

    def validateAndSetEnable(self, enable) -> bool:
        if enable in (0, 1, True, False):
            self.enable = enable
            return True
        return False

    def validateAndSetDrive(self, drive) -> bool:
        if drive in (0, 1, 2, 3):
            self.drive = drive
            return True
        return False

    def validateAndSetBoost(self, boost) -> bool:
        if boost in range(8):
            self.boost = boost
            return True
        return False


def read_value(address, bit, size):
    return 0


class GpioPortTableModel(QAbstractTableModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.items: List[GpioPort] = []
        self.readRegisterInfoes("register.xlsx")

    def readRegisterInfoes(self, filename):
        wb = openpyxl.load_workbook(filename, data_only=True)
        ws = wb['Register']
        for ir, row in enumerate(ws.iter_rows()):
            if ir > 0:
                gpioPort = GpioPort(
                    name=ws.cell(ir+1, 2).value, 
                    port=int(ws.cell(ir+1, 3).value,16),
                    signal_port = Register(int(ws.cell(ir+1, 5).value,16), ws.cell(ir+1, 6).value, ws.cell(ir+1, 7).value),
                    direction_port = Register(int(ws.cell(ir+1, 17).value,16), ws.cell(ir+1, 18).value, ws.cell(ir+1, 19).value),
                    value_port = Register(int(ws.cell(ir+1, 21).value,16), ws.cell(ir+1, 22).value, ws.cell(ir+1, 23).value),
                    enable_port = Register(int(ws.cell(ir+1, 25).value,16), ws.cell(ir+1, 26).value, ws.cell(ir+1, 27).value),
                    drive_port = Register(int(ws.cell(ir+1, 29).value,16), ws.cell(ir+1, 30).value, ws.cell(ir+1, 31).value),
                    boost_port = Register(int(ws.cell(ir+1, 33).value,16), ws.cell(ir+1, 34).value, ws.cell(ir+1, 35).value),
                    values_signal = [ws.cell(ir+1,8).value,ws.cell(ir+1,9).value,ws.cell(ir+1,10).value,
                                       ws.cell(ir+1,11).value,ws.cell(ir+1,12).value,ws.cell(ir+1,13).value,
                                       ws.cell(ir+1,14).value,ws.cell(ir+1,15).value],
                    signal = ws.cell(ir+1, 4).value,
                    direction = ws.cell(ir+1, 16).value,
                    value = ws.cell(ir+1, 20).value,
                    enable = ws.cell(ir+1, 24).value,
                    drive = ws.cell(ir+1, 28).value,
                    boost = ws.cell(ir+1, 32).value
                )
                gpioPort.signal = read_value(gpioPort.signal_port.address, gpioPort.signal_port.bit, gpioPort.signal_port.size)
                gpioPort.direction = read_value(gpioPort.direction_port.address, gpioPort.direction_port.bit, gpioPort.direction_port.size)
                gpioPort.value = read_value(gpioPort.value_port.address, gpioPort.value_port.bit, gpioPort.value_port.size)
                gpioPort.enable = read_value(gpioPort.enable_port.address, gpioPort.enable_port.bit, gpioPort.enable_port.size)
                gpioPort.drive = read_value(gpioPort.drive_port.address, gpioPort.drive_port.bit, gpioPort.drive_port.size)
                gpioPort.boost = read_value(gpioPort.boost_port.address, gpioPort.boost_port.bit, gpioPort.boost_port.size)
                self.items.append(gpioPort)


    def data(self, index: QModelIndex, role: int) -> Any:
        if role == Qt.DisplayRole:
            # The QTableView wants a cell text of 'index'
            # BE CAREFUL about IndexError (rowCount() and/or columnCount() are incorrect.)
            return self.items[index.row()].toList()[index.column()]

    def rowCount(self, parent=QModelIndex()) -> int:
        # = data count
        return len(self.items)

    def columnCount(self, parent=QModelIndex()) -> int:
        return len(GpioPort.toHeaderList())

    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.DisplayRole):
        if role == Qt.DisplayRole:
            # The QTableView wants a header text

            if orientation == Qt.Horizontal:
                # BE CAREFUL about IndexError
                return GpioPort.toHeaderList()[section]

            return ""  # There is no vertical header

    def setData(self, index: QModelIndex, value, role: int) -> bool:
        # Returning true means the value was accepted.
        if index.isValid() and role == Qt.EditRole:
            if index.column() == 0:
                return self.items[index.row()].validateAndSetName(value)
            elif index.column() == 1:
                return self.items[index.row()].validateAndSetPort(value)
            elif index.column() == 2:
                return self.items[index.row()].validateAndSetSignal(value)
            elif index.column() == 3:
                value = 0 if value==False else 1
                return self.items[index.row()].validateAndSetDirection(value)
            elif index.column() == 4:
                value = 0 if value==False else 1
                return self.items[index.row()].validateAndSetValue(value)
            elif index.column() == 5:
                value = 0 if value==False else 1
                return self.items[index.row()].validateAndSetEnable(value)
            elif index.column() == 6:
                return self.items[index.row()].validateAndSetDrive(value)
            elif index.column() == 7:
                value = 0 if value==False else 1
                return self.items[index.row()].validateAndSetBoost(value)
        return False  # Not Accepted.

    def flags(self, index: QModelIndex) -> Qt.ItemFlags:
        if index.isValid():
            return Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable

        return Qt.NoItemFlags


class MyComboBox(QComboBox):
    def __init__(self, parent, choices):
        super().__init__(parent)
        self.setFrame(False)
        self.setAutoFillBackground(True)
        for i, choice in enumerate(choices):
            self.insertItem(i, choice)
        
class MyOptionButton(QStyleOptionButton):
    def __init__(self, values, state):
        super().__init__()
        self.state |= QtWidgets.QStyle.State_Enabled
        if state == True:
            self.state |= QtWidgets.QStyle.State_On
        else:
            self.state |= QtWidgets.QStyle.State_Off
        self.text = values[0] if state==False else values[1]


class GpioRegisterTableDelegate(QtWidgets.QItemDelegate):
 
    def __init__(self, parent=None):
        self.parent = parent
        self.event_type = None
        super(GpioRegisterTableDelegate, self).__init__(parent)
 
    def createEditor(self, parent, option, index):
        """
        編集したいCellに対して、編集用のWidgetsを作成する
        """
        if index.column() == 0:
            return QtWidgets.QLineEdit(parent)
 
        elif index.column() == 8:
            spin = QtWidgets.QSpinBox(parent)
            spin.setMinimum(0)
            spin.setMaximum(100)
            return spin

        elif index.column() == self.parent.column_signal:
            values = self.parent.gpioPortTableModel.items[index.row()].values_signal
            combo = MyComboBox(parent, values)
            return combo

        elif index.column() == self.parent.column_drive:
            combo = MyComboBox(parent, self.parent.values_drive)
            return combo

    def setEditorData(self, editor, index):
        """
        createEditorで作成したWidgetsを受け取って、
        Editorにデフォルト値を設定する。
        今回の場合、元々のCellに表示されていた内容を受け取り、
        QLineEditに対してデフォルト値をセットしている
        """
 
        if index.column() == 0:
            value = index.model().data(index, QtCore.Qt.DisplayRole)
            editor.setText(value)
 
        elif index.column() == 8:
            value = index.model().data(index, QtCore.Qt.DisplayRole)
            editor.setValue(value)

        elif index.column() in [self.parent.column_signal, self.parent.column_drive]:
            value = index.model().data(index, QtCore.Qt.DisplayRole)
            editor.itemData(value)

    def editorEvent(self, event, model, option, index):
        """
        TableのCellに対してなにかしらのイベント（クリックした等）が発生したときに呼ばれる。
        """
        self.event_type = event.type() if self.event_type != QtCore.QEvent.MouseButtonDblClick else self.event_type
        if index.column() in [self.parent.column_direction, self.parent.column_value,
                              self.parent.column_enable, self.parent.column_boost]:
            if self.checkBoxRect(option).contains(event.pos().x(), event.pos().y()):
                if event.type() == QtCore.QEvent.MouseButtonPress:
                    currentValue = model.data(index, Qt.DisplayRole)
                    model.setData(index, not currentValue, Qt.EditRole)
                    model.layoutChanged.emit()
                    return True

        return False
 
    def setModelData(self, editor, model, index):
        """
        編集した値を、Modelを経由して実体のオブジェクトに対してセットする
        """
        value = None
        if index.column() == 0:
            value = editor.text()
        elif index.column() == 8:
            value = editor.value()
        elif index.column() in [self.parent.column_signal, self.parent.column_drive]:
            value = editor.currentText()
            value = int(value[0])
        else:
            value = editor.value()
            value = int(value[0])

        if value is not None:
             model.setData(index, value, Qt.EditRole)

        if index.column() == self.parent.column_signal:
            if (self.event_type == QtCore.QEvent.MouseButtonDblClick):
                    print(index.row(),index.column(), index.data())
                    self.event_type = None
 
    def checkBoxRect(self, option):
        rect = option.rect
        rect.setX(rect.x() + 10)
        rect.setWidth(rect.width() - 20)
        return rect
 
    def paint(self, painter, option, index):
        """
        Cellの中の描画を行う
        """
        # print(index.column(), index.data(), type(index.data()))
        col = index.column()
        if col == 0:
            painter.drawText(option.rect, QtCore.Qt.AlignCenter | QtCore.Qt.TextWordWrap, index.data())
        elif col == self.parent.column_port:
            txt = f'0x{index.data():02X}'
            painter.drawText(option.rect, QtCore.Qt.AlignCenter | QtCore.Qt.TextWordWrap, txt)
            
        elif col == 8:
            bar = QtWidgets.QStyleOptionProgressBar()
            bar.rect = option.rect
            bar.rect.setHeight(option.rect.height() - 1)
            bar.rect.setTop(option.rect.top() + 1)
            bar.minimum = 0
            bar.maximum = 100
            bar.progress = int(index.data())
            bar.textVisible = True
            bar.text = str(index.data()) + '%'
            bar.textAlignment = QtCore.Qt.AlignCenter
            QtWidgets.QApplication.style().drawControl(QtWidgets.QStyle.CE_ProgressBar, bar, painter)

        elif col == self.parent.column_signal:
            txt = self.parent.gpioPortTableModel.items[index.row()].values_signal[index.data()]
            painter.drawText(option.rect, QtCore.Qt.AlignCenter | QtCore.Qt.TextWordWrap, txt)

        elif col == self.parent.column_drive:
            txt = self.parent.values_drive[index.data()]
            painter.drawText(option.rect, QtCore.Qt.AlignCenter | QtCore.Qt.TextWordWrap, txt)

        # check button
        elif col in [self.parent.column_direction, self.parent.column_value,
                              self.parent.column_enable, self.parent.column_boost]:
            if index.column() == self.parent.column_direction:
                values = self.parent.values_direction
            elif index.column() == self.parent.column_value:
                values = self.parent.values_value
            elif index.column() == self.parent.column_enable:
                values = self.parent.values_enable
            elif index.column() == self.parent.column_boost:
                values = self.parent.values_boost
            state = False if index.data() == 0 else True
            btn = MyOptionButton(values, state)
            btn.rect = self.checkBoxRect(option)
            QtWidgets.QApplication.style().drawControl(QtWidgets.QStyle.CE_CheckBox, btn, painter)

        else:
            painter.drawText(option.rect, QtCore.Qt.AlignCenter | QtCore.Qt.TextWordWrap, str(index.data()))


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        if not MainWindow.objectName():
            MainWindow.setObjectName(u"MainWindow")
        MainWindow.resize(640, 696)
        sizePolicy = QSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(MainWindow.sizePolicy().hasHeightForWidth())
        MainWindow.setSizePolicy(sizePolicy)
        self.centralwidget = QWidget(MainWindow)
        self.centralwidget.setObjectName(u"centralwidget")
        self.verticalLayout = QVBoxLayout(self.centralwidget)
        self.verticalLayout.setObjectName(u"verticalLayout")
        self.tableView = QTableView(self.centralwidget)
        self.tableView.setObjectName(u"tableView")

        self.verticalLayout.addWidget(self.tableView)

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QMenuBar(MainWindow)
        self.menubar.setObjectName(u"menubar")
        self.menubar.setGeometry(QRect(0, 0, 640, 21))
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QStatusBar(MainWindow)
        self.statusbar.setObjectName(u"statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)

        QMetaObject.connectSlotsByName(MainWindow)
    # setupUi

    def retranslateUi(self, MainWindow):
        MainWindow.setWindowTitle(QCoreApplication.translate("MainWindow", u"MainWindow", None))
    # retranslateUi
        self.column_name = 0
        self.column_port = 1
        self.column_signal = 2
        self.column_direction = 3
        self.column_value = 4
        self.column_enable = 5
        self.column_drive = 6
        self.column_boost = 7
        self.values_value = ["Low", "High"]
        self.values_direction = ["Output", "Input"]
        self.values_enable = ["Disable", "Enable"]
        self.values_boost = ["Off", "On"]
        self.values_drive = ["0:Low", "1:Middle-Low", "2:Middle-High", "3:High"]
        self.values_signal = ["0:", "1:", "2:", "3:", "4:", "5:", "6:", "7:"]


class TestWindow(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.root_widget: QWidget = QWidget()
        self.layout: QVBoxLayout = QVBoxLayout()

        self.setupUi(self)

        self.gpioPortTableModel = GpioPortTableModel()

        self.tableView.setModel(self.gpioPortTableModel)  # create model and set
        
        self.delegate = GpioRegisterTableDelegate(self)

        #self.tableView.CurrentChanged.connect(self.delegate.editorEvent)

        self.tableView.setItemDelegate(self.delegate)
        print()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = TestWindow()
    window.show()

    sys.exit(app.exec_())